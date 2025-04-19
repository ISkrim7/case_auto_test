from typing import List, Dict, Any

from openpyxl import Workbook, load_workbook
from app.mapper.project.moduleMapper import ModuleMapper
from utils import GenerateTools


class ExcelClient:

    def init_api_excel(self):
        """
        初始化 api 导入模板
        :return:
        """
        work_book = Workbook()
        try:
            work_book.remove(work_book["Sheet"])  # 删除默认sheet
            # 初始化api sheet
            self._init_api(work_book)
            self._init_module(work_book)
            work_book.save(f"导入模板_{GenerateTools.getTime(6)}.xlsx")
        finally:
            work_book.close()  # 显式关闭

    def _init_api(self, wb: Workbook):
        """
        api sheet
        :param wb:
        :return:
        """
        sheet_api = wb.create_sheet("api")
        sheet_api_title = ["接口名", "接口描述", "接口状态", "接口等级", "请求方法", "接口地址", "请求参数", "请求体"]
        sheet_api.append(sheet_api_title)
        self._add_style(sheet_api)

    def _init_module(self, wb: Workbook):
        """
        module sheet
        :param wb:
        :return:
        """
        sheet_module = wb.create_sheet("模块")
        sheet_module_title = ["模块名", "模块ID"]
        sheet_module.append(sheet_module_title)
        self._add_style(sheet_module)

    @staticmethod
    def _add_style(sheet):
        """普通模式下设置样式"""
        from openpyxl.styles import Font, Alignment, PatternFill
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")  # 白色加粗
        header_fill = PatternFill(start_color="4F81BD", fill_type="solid")  # 蓝色背景
        header_alignment = Alignment(horizontal="center")  # 居中

        # 应用样式到表头行
        for cell in sheet[1]:  # ws[1] 表示第一行
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment


if __name__ == '__main__':
    ExcelClient().init_api_excel()
